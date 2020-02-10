import 	openpyxl
from    os import remove
from 	datetime import datetime
from 	datetime import timedelta
from 	os import listdir
from 	os.path import isfile, isdir

def ls1(path):    
    return [obj for obj in listdir(path) if isfile(path + obj)]


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):

            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def pasteRange1(startCol, startRow, endCol, endRow, sheetReceiving,valor):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):

            sheetReceiving.cell(row = i, column = j).value = valor
            countCol += 1
        countRow += 1  

def createData(ruta1,ruta2,file):
    print("Processing...")
    wb = openpyxl.load_workbook(ruta1) #Add file name
    sheet = wb['OC'] #Add Sheet name    
    #File to be pasted into
    template = openpyxl.load_workbook(ruta2) #Add file name
    temp_sheet = template['Hoja1'] #Add Sheet name
    max_row=sheet.max_row   
    print(max_row)
    max_row= max_row -1
    print(max_row)
    max_row1 = max_row - 16
    print(max_row1)
    cadena = sheet.cell(max_row, 6).value
    print(cadena)
    cadena1 = cadena[0:3]
    print(cadena1)
    now = datetime.now()
    new_date = now + timedelta(days=15)
    print(new_date)
    date_time = new_date.strftime("%d.%m.%Y")
    print(date_time)
    #You can save the template as another file to create a new file here too.s
    #template.save("ABC.xlsx")
    selectedRange = copyRange(3,16,3,max_row,sheet) #Change the 4 number values
    pastingRange  = pasteRange(14,2,14,2+max_row1,temp_sheet,selectedRange) #Change the 4 number values

    selectedRange = copyRange(7,16,7,max_row,sheet) #Change the 4 number values
    pastingRange = pasteRange(15,2,15,2+max_row1,temp_sheet,selectedRange) #Change the 4 number values

    selectedRange = copyRange(8,16,8,max_row,sheet) #Change the 4 number values
    pastingRange = pasteRange(16,2,16,2+max_row1,temp_sheet,selectedRange) #Change the 4 number values

 
    pastingRange = pasteRange1(1,2,1,2+max_row1,temp_sheet,'Z011') #Change the 4 number values
    pastingRange = pasteRange1(2,2,2,2+max_row1,temp_sheet,'2000') #Change the 4 number values
    pastingRange = pasteRange1(3,2,3,2+max_row1,temp_sheet,'30') #Change the 4 number values
    pastingRange = pasteRange1(4,2,4,2+max_row1,temp_sheet,'00') #Change the 4 number values
    pastingRange = pasteRange1(5,2,5,2+max_row1,temp_sheet,'1102') #Change the 4 number values
    pastingRange = pasteRange1(6,2,6,2+max_row1,temp_sheet,'010') #Change the 4 number values
    pastingRange = pasteRange1(7,2,7,2+max_row1,temp_sheet,'1029206') #Change the 4 number values
    #pastingRange = pasteRange1(8,2,8,2+max_row1,temp_sheet,cadena1) #Change the 4 number values
    pastingRange = pasteRange1(8,2,8,2+max_row1,temp_sheet,'504') #Change the 4 number values para que funcione    
    pastingRange = pasteRange1(9,2,9,2+max_row1,temp_sheet,'REGULARIZACION') #Change the 4 number values
    pastingRange = pasteRange1(17,2,17,2+max_row1,temp_sheet,967849) #Change the 4 number values
    pastingRange = pasteRange1(18,2,18,2+max_row1,temp_sheet,date_time) #Change the 4 number values

    template.save("C:\\Users\\canviademo\\Desktop\\celima\\Corporacion Celima\\ZSDP580\\"+file+"ZSDP580"+".xlsx")
    print("Range copied and pasted!")

				
if __name__ == "__main__":
    arvhivos = [] 
    ruta = "C:\\Users\\canviademo\\Desktop\\celima\\Corporacion Celima\\"
    arvhivos=ls1(ruta)
    print(arvhivos)
    for v in arvhivos:
        print("C:\\Users\\canviademo\\Desktop\\celima\\Corporacion Celima\\"+v)
        print(v.find(".xlsx"))
        if v.find(".xlsx") == -1:
            pass
        else:
            print("b")   
            print(len(v)-4)
            file = v[0:len(v)-5]
            print(file)
            createData("C:\\Users\\canviademo\\Desktop\\celima\\Corporacion Celima\\"+v,"ZSDP580 - copia.xlsx",file)
            remove("C:\\Users\\canviademo\\Desktop\\celima\\Corporacion Celima\\"+v)